"""
Cancellation guidance for detected recurring services.

The app uses a small curated database for common services and falls back to a
vendor-specific web-search link when a service is unknown.
"""

from __future__ import annotations

import os
import re
import json
from dataclasses import dataclass, field
from datetime import date
from functools import lru_cache
from pathlib import Path
from typing import Any
from urllib.parse import quote_plus
from urllib.request import Request, urlopen
from urllib.error import HTTPError, URLError

from merchant_normalizer import clean


@dataclass
class CancellationGuide:
    service_name: str
    matched_name: str
    confidence: float
    source: str
    manage_url: str
    source_url: str
    search_url: str
    category: str = ""
    market_position: str = ""
    price_range: str = ""
    billing_cycle: str = ""
    website: str = ""
    cancellation_process: str = ""
    additional_resources: str = ""
    database_path: str = ""
    steps: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def to_dict(self) -> dict:
        return {
            "service_name": self.service_name,
            "matched_name": self.matched_name,
            "confidence": self.confidence,
            "source": self.source,
            "manage_url": self.manage_url,
            "source_url": self.source_url,
            "search_url": self.search_url,
            "category": self.category,
            "market_position": self.market_position,
            "price_range": self.price_range,
            "billing_cycle": self.billing_cycle,
            "website": self.website,
            "cancellation_process": self.cancellation_process,
            "additional_resources": self.additional_resources,
            "database_path": self.database_path,
            "steps": self.steps,
            "notes": self.notes,
        }


@dataclass(frozen=True)
class _GuideRecord:
    service_name: str
    aliases: tuple[str, ...]
    manage_url: str
    source_url: str
    steps: tuple[str, ...]
    notes: tuple[str, ...] = ()


@dataclass(frozen=True)
class _WorkbookGuideRecord:
    service_name: str
    category: str
    market_position: str
    price_range: str
    billing_cycle: str
    website: str
    cancellation_process: str
    additional_resources: str
    database_path: str
    aliases: tuple[str, ...]


_DEFAULT_DATABASE_PATHS = (
    Path(__file__).with_name("subscription_cancellation_process.xlsx"),
    Path.home() / "Downloads" / "subscription_cancellation_process.xlsx",
)
_OPENAI_RESPONSES_URL = "https://api.openai.com/v1/responses"
_OPENAI_MODEL_ENV = "OPENAI_WEBSEARCH_MODEL"
_OPENAI_DEFAULT_MODEL = "gpt-4.1-mini"
_DATABASE_HEADERS = [
    "#",
    "Provider",
    "Market Position",
    "Price Range",
    "Billing Cycle",
    "Website",
    "Cancellation Process",
    "additional resources",
]
_LLM_SHEET_NAME = "LLM Web Search"


_GUIDES: tuple[_GuideRecord, ...] = (
    _GuideRecord(
        service_name="iCloud+",
        aliases=("icloud+", "icloud", "icloud plus"),
        manage_url="https://account.apple.com/",
        source_url="https://support.apple.com/en-us/118428",
        steps=(
            "On iPhone or iPad, open Settings, tap your name, then tap Subscriptions.",
            "Select iCloud+ or iCloud, then tap Cancel Subscription or downgrade the storage plan.",
            "On Mac, open App Store, click your name, open Account Settings, then manage Subscriptions.",
            "Without an Apple device, sign in at account.apple.com or use reportaproblem.apple.com for recent iCloud+ charges.",
        ),
        notes=(
            "Use the same Apple Account shown on the receipt.",
            "If the subscription belongs to a family member, that account owner has to cancel it.",
        ),
    ),
    _GuideRecord(
        service_name="Apple / App Store",
        aliases=("apple", "apple com bill", "itunes", "app store"),
        manage_url="https://account.apple.com/",
        source_url="https://support.apple.com/en-us/118428",
        steps=(
            "On iPhone or iPad, open Settings, tap your name, then tap Subscriptions.",
            "Choose the subscription and tap Cancel Subscription.",
            "On Mac, open App Store, click your name, open Account Settings, then manage Subscriptions.",
            "On Windows, use Apple Music, Apple TV, or iTunes account settings to manage subscriptions.",
        ),
        notes=(
            "If there is no Cancel button, Apple says the subscription is already canceled.",
            "If the charge was billed by another company, cancel through that biller instead.",
        ),
    ),
    _GuideRecord(
        service_name="Netflix",
        aliases=("netflix", "nflx", "netflix com"),
        manage_url="https://www.netflix.com/cancelplan",
        source_url="https://help.netflix.com/en/node/407",
        steps=(
            "Sign in to Netflix and open the cancel or membership management page.",
            "Select Cancel.",
            "Select Finish Cancellation and watch for the confirmation email.",
        ),
        notes=(
            "Signing out or deleting the app does not cancel Netflix.",
            "If Netflix says a payment partner manages billing, cancel through that partner.",
        ),
    ),
    _GuideRecord(
        service_name="Spotify",
        aliases=("spotify", "spotify premium", "spotify usa"),
        manage_url="https://www.spotify.com/account/subscription/",
        source_url="https://support.spotify.com/us/article/cancel-premium/",
        steps=(
            "Sign in to Spotify and open Manage your plan.",
            "Select Cancel subscription.",
            "Confirm the cancellation flow; Premium stays active until the next billing date.",
        ),
        notes=(
            "If there is no cancel option, check the Payment section for the partner company that bills you.",
            "Plan members in Duo or Family plans need the plan manager to cancel the paid plan.",
        ),
    ),
    _GuideRecord(
        service_name="Adobe",
        aliases=("adobe", "adobe creative", "adobe creative cloud"),
        manage_url="https://account.adobe.com/plans",
        source_url="https://helpx.adobe.com/account/individual/subscriptions-and-plans/renewals-and-cancellations/cancel-adobe-subscription.html",
        steps=(
            "Sign in to your Adobe account plans page.",
            "Select Manage plan for the plan you want to cancel.",
            "Select Cancel your plan, continue through the reason and review screens, then confirm.",
        ),
        notes=(
            "If you purchased through Apple, Google, or Microsoft, cancel through that provider.",
            "Adobe refund eligibility depends on the plan and when you cancel.",
        ),
    ),
    _GuideRecord(
        service_name="Canva",
        aliases=("canva", "canva pro", "canva teams"),
        manage_url="https://www.canva.com/settings/billing",
        source_url="https://www.canva.com/help/cancel-canva-plan/",
        steps=(
            "Sign in to the Canva account and team that owns the paid plan.",
            "Open Settings, then Billing and plans.",
            "Find the active plan and choose Cancel plan or Manage subscription, then confirm.",
        ),
        notes=(
            "Cancel through Apple App Store or Google Play if that is where the subscription was purchased.",
            "Deleting the Canva app does not cancel the subscription.",
        ),
    ),
    _GuideRecord(
        service_name="Microsoft 365",
        aliases=("microsoft", "microsoft 365", "office365", "office 365", "onedrive"),
        manage_url="https://account.microsoft.com/services/microsoft365",
        source_url="https://support.microsoft.com/en-us/accounts-billing/subscriptions/cancel-a-microsoft-365-subscription",
        steps=(
            "Sign in at Microsoft services with the account used to buy Microsoft 365.",
            "Select Cancel subscription, Upgrade or Cancel, or turn off recurring billing.",
            "Review the cancellation page and choose the option that confirms you do not want the subscription.",
        ),
        notes=(
            "If purchased from Apple, Google Play, Amazon, or another retailer, cancel through that seller.",
            "Microsoft warns that storage allowances may drop after cancellation.",
        ),
    ),
    _GuideRecord(
        service_name="YouTube Premium",
        aliases=("youtube", "youtube premium", "youtube music", "google youtube"),
        manage_url="https://www.youtube.com/paid_memberships",
        source_url="https://support.google.com/youtube/answer/6308278",
        steps=(
            "Open youtube.com/paid_memberships while signed in.",
            "Click Manage membership, then Deactivate.",
            "Choose Continue to cancel, select a reason, then confirm with Yes, cancel.",
        ),
        notes=(
            "If billed by Apple or Google Play, cancel through that subscription store.",
            "Benefits continue until the end of the billing period after cancellation.",
        ),
    ),
)


class WebSearchCancellationError(RuntimeError):
    """Raised when API-backed cancellation research cannot complete."""


def get_cancellation_guide(merchant_raw: str, *, category: str | None = None) -> CancellationGuide:
    """Return the best cancellation guide for a merchant or a web-search fallback."""
    merchant = (merchant_raw or "").strip()
    cleaned = clean(merchant)
    search_url = _search_url(merchant or "subscription")

    best: tuple[object, str, float, int] | None = None
    for record, priority in [*[(r, 2) for r in _load_workbook_guides()], *[(r, 1) for r in _GUIDES]]:
        for alias in record.aliases:
            score = _match_score(cleaned, clean(alias))
            if score <= 0:
                continue
            if best is None or score > best[2] or (score == best[2] and priority > best[3]):
                best = (record, alias, score, priority)

    if best is not None:
        record, alias, score, _priority = best
        if isinstance(record, _WorkbookGuideRecord):
            return _workbook_record_to_guide(record, alias, score, search_url)
        if isinstance(record, _GuideRecord):
            return _built_in_record_to_guide(record, alias, score, search_url)

    label = merchant.title() if merchant else "Unknown subscription"
    category_note = f"Detected category: {category}." if category else "No known cancellation profile found yet."
    process = "\n".join(
        [
            "Type: Web-search fallback",
            "Use when this provider is not in the local cancellation database.",
            f"Web search: {search_url}",
            "",
            "Workflow:",
            "1. Open the vendor's official account, billing, plan, or subscriptions page.",
            "2. Look for Manage plan, Billing, Subscriptions, Cancel, or Turn off recurring billing.",
            "3. Confirm cancellation and keep the confirmation email or screenshot.",
            "",
            "Watch-out:",
            "Verify the domain is the provider's official site before signing in.",
        ]
    )
    return CancellationGuide(
        service_name=label,
        matched_name=merchant,
        confidence=0.0,
        source="web_search",
        manage_url="",
        source_url="",
        search_url=search_url,
        category=category or "",
        cancellation_process=process,
        additional_resources=search_url,
        steps=[
            "Open the vendor's official account, billing, plan, or subscriptions page.",
            "Look for Manage plan, Billing, Subscriptions, Cancel, or Turn off recurring billing.",
            "Confirm cancellation and keep the confirmation email or screenshot.",
        ],
        notes=[
            category_note,
            "Use the web-search link to find current official cancellation instructions before entering credentials.",
        ],
    )


def research_and_update_cancellation_guide(
    merchant_raw: str,
    *,
    category: str | None = None,
    database_path: str | Path | None = None,
) -> CancellationGuide:
    """
    Use the OpenAI API with web search to research an unknown cancellation flow,
    append/update the Excel cancellation database, and return the saved guide.

    Requires OPENAI_API_KEY. Optionally set OPENAI_WEBSEARCH_MODEL.
    """
    merchant = (merchant_raw or "").strip()
    if not merchant:
        raise WebSearchCancellationError("A merchant/provider name is required.")

    existing = get_cancellation_guide(merchant, category=category)
    if existing.source == "xlsx_database":
        return existing

    researched = _research_with_openai(merchant, category=category)
    path = save_researched_cancellation_record(researched, database_path=database_path)
    clear_cancellation_database_cache()

    saved = get_cancellation_guide(researched["provider"], category=category)
    if saved.source == "xlsx_database":
        return saved

    # If matching by returned provider name fails, still return the researched
    # content in the app's normal shape.
    return _researched_dict_to_guide(researched, str(path))


def save_researched_cancellation_record(
    record: dict[str, Any],
    *,
    database_path: str | Path | None = None,
) -> Path:
    """Append or update one researched provider in the Excel cancellation database."""
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError as exc:
        raise WebSearchCancellationError("openpyxl is required to update the cancellation workbook.") from exc

    path = Path(database_path) if database_path else cancellation_database_path()
    if path is None:
        path = Path(__file__).with_name("subscription_cancellation_process.xlsx")

    if path.exists():
        workbook = load_workbook(path)
    else:
        workbook = Workbook()
        workbook.active.title = _LLM_SHEET_NAME

    sheet = _ensure_llm_sheet(workbook)
    _upsert_sheet_row(sheet, _normalize_researched_record(record))
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    clear_cancellation_database_cache()
    return path


def clear_cancellation_database_cache() -> None:
    _load_workbook_guides.cache_clear()


def cancellation_database_path() -> Path | None:
    """Return the workbook path used as the cancellation database, if available."""
    env_path = os.environ.get("SUBSCRIPTION_CANCELLATION_XLSX")
    candidates = [Path(env_path)] if env_path else []
    candidates.extend(_DEFAULT_DATABASE_PATHS)
    for candidate in candidates:
        if candidate and candidate.exists():
            return candidate
    return None


@lru_cache(maxsize=1)
def _load_workbook_guides() -> tuple[_WorkbookGuideRecord, ...]:
    path = cancellation_database_path()
    if path is None:
        return ()

    try:
        from openpyxl import load_workbook
    except ImportError:
        return ()

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return ()

    records: list[_WorkbookGuideRecord] = []
    for sheet in workbook.worksheets:
        if sheet.title.lower() == "index":
            continue
        rows = list(sheet.iter_rows(values_only=True))
        header_idx, header_map = _find_header(rows)
        if header_idx is None:
            continue

        provider_idx = header_map["provider"]
        cancel_idx = header_map["cancellation"]
        for row in rows[header_idx + 1 :]:
            provider = _cell(row, provider_idx)
            process = _cell(row, cancel_idx)
            if not provider or not process:
                continue
            if provider.lower() == "provider" or provider.startswith("#") or provider.startswith("Warning"):
                continue
            records.append(
                _WorkbookGuideRecord(
                    service_name=provider,
                    category=sheet.title,
                    market_position=_cell(row, header_map.get("market_position")),
                    price_range=_cell(row, header_map.get("price_range")),
                    billing_cycle=_cell(row, header_map.get("billing_cycle")),
                    website=_cell(row, header_map.get("website")),
                    cancellation_process=process,
                    additional_resources=_cell(row, header_map.get("resources")),
                    database_path=str(path),
                    aliases=_provider_aliases(provider),
                )
            )
    return tuple(records)


def _research_with_openai(merchant: str, *, category: str | None = None) -> dict[str, Any]:
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise WebSearchCancellationError(
            "Set OPENAI_API_KEY before researching unknown cancellation policies."
        )

    model = os.environ.get(_OPENAI_MODEL_ENV, _OPENAI_DEFAULT_MODEL).strip() or _OPENAI_DEFAULT_MODEL
    prompt = _research_prompt(merchant, category=category)
    payload = {
        "model": model,
        "tools": [{"type": "web_search", "search_context_size": "medium"}],
        "tool_choice": "required",
        "input": prompt,
        "max_output_tokens": 1800,
    }
    data = json.dumps(payload).encode("utf-8")
    request = Request(
        _OPENAI_RESPONSES_URL,
        data=data,
        headers={
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
        },
        method="POST",
    )
    try:
        with urlopen(request, timeout=60) as response:
            raw = response.read().decode("utf-8")
    except HTTPError as exc:
        detail = exc.read().decode("utf-8", errors="replace")
        raise WebSearchCancellationError(f"OpenAI API request failed: {exc.code} {detail}") from exc
    except URLError as exc:
        raise WebSearchCancellationError(f"OpenAI API request failed: {exc.reason}") from exc
    except TimeoutError as exc:
        raise WebSearchCancellationError("OpenAI API request timed out.") from exc

    try:
        response_payload = json.loads(raw)
    except json.JSONDecodeError as exc:
        raise WebSearchCancellationError("OpenAI API returned non-JSON response.") from exc

    text = _extract_response_text(response_payload)
    if not text:
        raise WebSearchCancellationError("OpenAI API returned no cancellation research text.")
    parsed = _parse_research_json(text)
    return _normalize_researched_record(parsed)


def _research_prompt(merchant: str, *, category: str | None = None) -> str:
    category_hint = category or "unknown"
    return f"""
Research current official cancellation instructions for this recurring subscription provider:

Provider: {merchant}
Detected category: {category_hint}

Use web search. Prefer official provider support, account, billing, help, app-store, or regulator pages.
If the provider is billed through Apple, Google Play, Amazon, Stripe, PayPal, a mobile carrier, or another partner,
include that as a watch-out. Do not invent support phone numbers, refund rules, or URLs.

Return ONLY a valid JSON object with exactly these keys:
{{
  "provider": "Provider name",
  "category": "Best matching database category or general",
  "market_position": "Short coverage/position note, or AI/web researched if unknown",
  "price_range": "Known price range or Unknown",
  "billing_cycle": "Monthly / Annual / Unknown etc.",
  "website": "provider domain or official account URL",
  "cancellation_process": "Workbook-style multiline text: Type 1, Direct link, Workflow numbered steps, Refund if known, Pause/freeze if known, Watch-out, Support if known",
  "additional_resources": "Official source URLs and titles, one per line"
}}

Keep the cancellation_process in the same style as an Excel cell in the existing database. Use concise but actionable text.
""".strip()


def _extract_response_text(payload: dict[str, Any]) -> str:
    chunks: list[str] = []
    for item in payload.get("output", []):
        if item.get("type") != "message":
            continue
        for content in item.get("content", []):
            if content.get("type") in {"output_text", "text"} and content.get("text"):
                chunks.append(str(content["text"]))
    if chunks:
        return "\n".join(chunks).strip()
    if payload.get("output_text"):
        return str(payload["output_text"]).strip()
    return ""


def _parse_research_json(text: str) -> dict[str, Any]:
    cleaned_text = text.strip()
    if cleaned_text.startswith("```"):
        cleaned_text = re.sub(r"^```(?:json)?\s*", "", cleaned_text, flags=re.IGNORECASE)
        cleaned_text = re.sub(r"\s*```$", "", cleaned_text)
    match = re.search(r"\{.*\}", cleaned_text, re.DOTALL)
    if match:
        cleaned_text = match.group(0)
    try:
        parsed = json.loads(cleaned_text)
    except json.JSONDecodeError as exc:
        raise WebSearchCancellationError("OpenAI response was not valid JSON in the required format.") from exc
    if not isinstance(parsed, dict):
        raise WebSearchCancellationError("OpenAI response must be a JSON object.")
    return parsed


def _normalize_researched_record(record: dict[str, Any]) -> dict[str, str]:
    normalized = {
        "provider": _text_field(record, "provider"),
        "category": _text_field(record, "category"),
        "market_position": _text_field(record, "market_position") or "AI/web researched",
        "price_range": _text_field(record, "price_range") or "Unknown",
        "billing_cycle": _text_field(record, "billing_cycle") or "Unknown",
        "website": _text_field(record, "website"),
        "cancellation_process": _text_field(record, "cancellation_process"),
        "additional_resources": _text_field(record, "additional_resources"),
    }
    if not normalized["provider"]:
        raise WebSearchCancellationError("Researched record is missing provider.")
    if not normalized["cancellation_process"]:
        raise WebSearchCancellationError("Researched record is missing cancellation_process.")
    if normalized["category"] and normalized["category"].lower() not in normalized["market_position"].lower():
        normalized["market_position"] = (
            f"{normalized['market_position']}\nCategory: {normalized['category']}"
        ).strip()
    return normalized


def _text_field(record: dict[str, Any], key: str) -> str:
    value = record.get(key, "")
    if isinstance(value, list):
        value = "\n".join(str(v) for v in value if v is not None)
    return re.sub(r"[ \t]+\n", "\n", str(value or "")).strip()


def _researched_dict_to_guide(record: dict[str, str], database_path: str) -> CancellationGuide:
    return CancellationGuide(
        service_name=record["provider"],
        matched_name=record["provider"],
        confidence=1.0,
        source="xlsx_database",
        manage_url=_first_url(record["cancellation_process"]) or _site_url(record["website"]),
        source_url=_first_url(record["additional_resources"]) or _first_url(record["cancellation_process"]),
        search_url=_search_url(record["provider"]),
        category=_LLM_SHEET_NAME,
        market_position=record["market_position"],
        price_range=record["price_range"],
        billing_cycle=record["billing_cycle"],
        website=record["website"],
        cancellation_process=record["cancellation_process"],
        additional_resources=record["additional_resources"],
        database_path=database_path,
        notes=[
            "Created by API-backed web search and saved to the cancellation workbook.",
            "Verify current terms on the official site before acting.",
        ],
    )


def _ensure_llm_sheet(workbook: Any) -> Any:
    if _LLM_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[_LLM_SHEET_NAME]
    else:
        sheet = workbook.create_sheet(_LLM_SHEET_NAME)

    header_row = _header_row_index(sheet)
    if header_row is None:
        sheet.delete_rows(1, sheet.max_row)
        sheet.append([f"{_LLM_SHEET_NAME} — API researched cancellation procedures"])
        sheet.append(["Generated entries follow the same columns as the source cancellation workbook."])
        sheet.append([])
        sheet.append(_DATABASE_HEADERS)
    return sheet


def _header_row_index(sheet: Any) -> int | None:
    for row in range(1, min(sheet.max_row, 25) + 1):
        values = [_normalize_header(sheet.cell(row=row, column=col).value) for col in range(1, 9)]
        if "provider" in values and any("cancellation" in v for v in values):
            return row
    return None


def _upsert_sheet_row(sheet: Any, record: dict[str, str]) -> None:
    header_row = _header_row_index(sheet)
    if header_row is None:
        sheet.append(_DATABASE_HEADERS)
        header_row = sheet.max_row

    provider_clean = clean(record["provider"])
    provider_col = 2
    target_row = None
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=provider_col).value
        if value and clean(str(value)) == provider_clean:
            target_row = row_idx
            break
    if target_row is None:
        target_row = sheet.max_row + 1

    row_values = [
        _next_row_number(sheet, header_row) if target_row > sheet.max_row else sheet.cell(target_row, 1).value,
        record["provider"],
        record["market_position"],
        record["price_range"],
        record["billing_cycle"],
        record["website"],
        record["cancellation_process"],
        record["additional_resources"],
    ]
    for col, value in enumerate(row_values, start=1):
        sheet.cell(row=target_row, column=col, value=value)


def _next_row_number(sheet: Any, header_row: int) -> int:
    max_num = 0
    for row_idx in range(header_row + 1, sheet.max_row + 1):
        value = sheet.cell(row=row_idx, column=1).value
        if isinstance(value, int):
            max_num = max(max_num, value)
    return max_num + 1


def _find_header(rows: list[tuple]) -> tuple[int | None, dict[str, int]]:
    for i, row in enumerate(rows[:25]):
        normalized = [_normalize_header(v) for v in row]
        if "provider" not in normalized:
            continue
        cancellation_cols = [
            idx for idx, value in enumerate(normalized) if "cancellation" in value
        ]
        if not cancellation_cols:
            continue

        return i, {
            "provider": normalized.index("provider"),
            "market_position": _header_index(normalized, "market position", "2026 market share"),
            "price_range": _header_index(normalized, "price range"),
            "billing_cycle": _header_index(normalized, "billing cycle"),
            "website": _header_index(normalized, "website"),
            "cancellation": cancellation_cols[0],
            "resources": _header_index(normalized, "additional resources", "perplexity cancellation process"),
        }
    return None, {}


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _header_index(headers: list[str], *names: str) -> int | None:
    for name in names:
        if name in headers:
            return headers.index(name)
    return None


def _cell(row: tuple, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    value = row[idx]
    if value is None:
        return ""
    return re.sub(r"[ \t]+\n", "\n", str(value)).strip()


def _provider_aliases(provider: str) -> tuple[str, ...]:
    aliases = {provider}
    no_parens = re.sub(r"\s*\([^)]*\)", "", provider).strip()
    if no_parens:
        aliases.add(no_parens)
    for piece in re.split(r"\s*/\s*|\s+-\s+|\s+\|\s+", provider):
        piece = piece.strip()
        if piece:
            aliases.add(piece)
    return tuple(sorted(aliases, key=len, reverse=True))


def _workbook_record_to_guide(
    record: _WorkbookGuideRecord,
    alias: str,
    score: float,
    search_url: str,
) -> CancellationGuide:
    manage_url = _first_url(record.cancellation_process) or _site_url(record.website)
    source_url = _first_url(record.additional_resources) or _first_url(record.cancellation_process) or manage_url
    return CancellationGuide(
        service_name=record.service_name,
        matched_name=alias,
        confidence=round(score, 2),
        source="xlsx_database",
        manage_url=manage_url,
        source_url=source_url,
        search_url=search_url,
        category=record.category,
        market_position=record.market_position,
        price_range=record.price_range,
        billing_cycle=record.billing_cycle,
        website=record.website,
        cancellation_process=record.cancellation_process,
        additional_resources=record.additional_resources,
        database_path=record.database_path,
        notes=[
            "Loaded from the subscription cancellation workbook.",
            "Verify current terms on the official site before acting.",
        ],
    )


def _built_in_record_to_guide(
    record: _GuideRecord,
    alias: str,
    score: float,
    search_url: str,
) -> CancellationGuide:
    process = "\n".join(
        [
            f"Type: {record.service_name}",
            f"Direct link: {record.manage_url}",
            "",
            "Workflow:",
            *[f"{i}. {step}" for i, step in enumerate(record.steps, 1)],
            "",
            "Watch-out:",
            *[f"- {note}" for note in record.notes],
        ]
    )
    return CancellationGuide(
        service_name=record.service_name,
        matched_name=alias,
        confidence=round(score, 2),
        source="built_in_database",
        manage_url=record.manage_url,
        source_url=record.source_url,
        search_url=search_url,
        website=record.manage_url,
        cancellation_process=process,
        additional_resources=record.source_url,
        steps=list(record.steps),
        notes=list(record.notes),
    )


def _match_score(cleaned: str, alias_clean: str) -> float:
    if not cleaned or not alias_clean:
        return 0.0
    if cleaned == alias_clean:
        return 1.0
    if alias_clean in cleaned:
        return 0.92
    if cleaned in alias_clean:
        return 0.86
    cleaned_tokens = set(cleaned.split())
    alias_tokens = set(alias_clean.split())
    if cleaned_tokens and alias_tokens and alias_tokens.issubset(cleaned_tokens):
        return 0.82
    return 0.0


def _search_url(merchant: str) -> str:
    query = f"{merchant} official cancel subscription"
    return f"https://www.google.com/search?q={quote_plus(query)}"


def _first_url(text: str) -> str:
    match = re.search(r"https?://[^\s)>\]]+", text or "")
    return match.group(0).rstrip(".,;") if match else ""


def _site_url(site: str) -> str:
    site = (site or "").strip()
    if not site:
        return ""
    if site.startswith(("http://", "https://")):
        return site
    if "." in site:
        return f"https://{site}"
    return ""
